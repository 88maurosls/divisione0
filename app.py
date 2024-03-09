import streamlit as st
import pandas as pd
import uuid
import re
from openpyxl import load_workbook
from openpyxl.styles import Font
import os
import gspread as gs
import glob


def correct_csv(file_path, expected_fields):
    with open(file_path, 'r', encoding='utf-8') as file:
        lines = file.readlines()

    corrected_lines = []
    for line in lines:
        fields = line.strip().split(',')
        if len(fields) > expected_fields:
            corrected_fields = fields[:expected_fields]
        else:
            corrected_fields = fields
        corrected_lines.append(','.join(corrected_fields) + '\n')

    with open(file_path, 'w', encoding='utf-8') as file:
        file.writelines(corrected_lines)

def trasponi_valore_accanto_header1(file_path, expected_fields,UPLOAD_FOLDER):
    try:
        # Assicurati che il file CSV sia nel formato corretto
        correct_csv(file_path, expected_fields)

        with open(file_path, 'r', encoding='utf-8') as file:
            lines = file.readlines()

        values_next_to_header1 = []
        data_rows = []
        header_pattern = re.compile(r'^HEADER\d+')
        value_next_to_header1 = ""

        for line in lines:
            fields = line.strip().split(',')
            if fields[0] == "HEADER1":
                value_next_to_header1 = fields[1] if len(fields) > 1 else ""
            if not header_pattern.match(fields[0]):
                data_rows.append(fields)
                values_next_to_header1.append(value_next_to_header1)

        df_data = pd.DataFrame(data_rows)

        # Gestione della colonna SKU
        if len(df_data.columns) >= 3:
            df_data['SKU'] = df_data.iloc[:, 1] + '-' + df_data.iloc[:, 2]

        df_data['Value_Next_to_HEADER1'] = values_next_to_header1
        df_data = df_data[['Value_Next_to_HEADER1', 'SKU'] + [col for col in df_data.columns if col not in ['SKU', 'Value_Next_to_HEADER1']]]

        # Genera un nome di file unico per evitare sovrascritture
        unique_filename = f"output_{uuid.uuid4()}.xlsx"
        temp_output_file_path = os.path.join(UPLOAD_FOLDER, unique_filename)

        headers = ["Rif. Sped.", "SKU", "Collo", "Codice", "Colore", "Size", "UPC", "Made in", "Unità", "Confezione", "customer PO", "Riferimento Spedizione"]
        df_data.columns = headers[:len(df_data.columns)]

        df_data[['Collo', 'UPC']] = df_data[['Collo', 'UPC']].apply(lambda x: '{:.0f}'.format(x) if isinstance(x, (int, float)) else x)

        df_data.to_excel(temp_output_file_path, index=False, float_format="%.0f")

        # Applica lo stile alle celle dell'intestazione
        wb = load_workbook(temp_output_file_path)
        ws = wb.active
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = Font(bold=True)
        wb.save(temp_output_file_path)

        return temp_output_file_path

    except Exception as e:
        print(f"Si è verificato un errore: {e}")
        return None

######################################
#### MAIN ############################
######################################
def main():

    st.title('App x Google Sheet')

    m = st.markdown("""
    <style>
    div.stButton > button:first-child {
        background-color: #0099ff;
        color:#ffffff;
    }
    div.stButton > button:hover {
        background-color: #00ff00;
        color:#ff0000;
        }
    </style>""", unsafe_allow_html=True)

    

    ####################################
    # parameters
    path = "uploads/Nike CSV-18e1a8c7be0.CSV"
    UPLOAD_FOLDER = "downloads"
    trasponi_valore_accanto_header1(path,9,UPLOAD_FOLDER)

    #####################################
    ##merging all Excel inside upload folders
    files = os.path.join("downloads/", "*.xlsx")
    files = glob.glob(files)
    #dfs = [pd.read_excel(file, dtype=column_data_types) for file in files]
    dfs = [pd.read_excel(file, ) for file in files]
    df = pd.concat(dfs, ignore_index=True)
    df.to_excel("partial_tot.xlsx",index=False)

    st.dataframe(df)

    if st.button('Publish G-sheet'):
        ######## append  to google sheet ######################
        #id=https://docs.google.com/spreadsheets/d/18mSCmOwv5k8on2M96v_TyTKiJVo5xbM-wwZyxIHLgRQ/edit#gid=0
        #condivedere il google sheet con la mail:"test-769@mygpt-416217.iam.gserviceaccount.com" ## quella del json per intenderci
        df = df.fillna('')
        gsheetId = '18mSCmOwv5k8on2M96v_TyTKiJVo5xbM-wwZyxIHLgRQ'
        gc = gs.service_account(filename="credential_gsheet.json")
        sh = gc.open_by_key(gsheetId)
        worksheet = sh.get_worksheet(0)#index sheet inside file
        #data_list = df.values.tolist() 
        #worksheet.append_rows(data_list)

        worksheet.clear() #clear sheet
        #replace all values
        worksheet.update([df.columns.values.tolist()] + df.values.tolist())
        st.write('Caricato su GoogleSheet!')
        st.balloons()
        st.write(f"check GoogleSheet at this [link](https://docs.google.com/spreadsheets/d/{gsheetId}/edit#gid=0)")

###### transformation #####################################

if __name__ == "__main__":
    main()
